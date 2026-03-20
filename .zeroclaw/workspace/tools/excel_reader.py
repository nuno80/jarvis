#!/usr/bin/env python3
"""
excel_reader.py — Tool per leggere file Excel (.xlsx/.xls) nel workspace ZeroClaw.

Uso:
    python3 excel_reader.py <file> [opzioni]

Opzioni:
    --sheet NOME        Foglio da leggere (default: primo foglio)
    --range A1:D10      Range celle in notazione A1 (default: tutto)
    --format FMT        Formato output: markdown|csv|json (default: markdown)
    --max-rows N        Numero massimo di righe (default: 100)
    --list-sheets       Elenca i fogli disponibili ed esci
    --header ROW        Riga da usare come intestazione, 1-indexed (default: 1)
    --no-header         Non usare intestazione, numera le colonne
"""

import argparse
import json
import os
import sys

MAX_OUTPUT_BYTES = 50_000  # 50 KB per evitare overflow di contesto


def parse_cell_ref(ref: str) -> tuple[int, int]:
    """Converte riferimento cella A1 -> (riga 0-indexed, colonna 0-indexed)."""
    ref = ref.strip().upper()
    col_str = ""
    row_str = ""
    for ch in ref:
        if ch.isalpha():
            col_str += ch
        else:
            row_str += ch

    if not col_str or not row_str:
        raise ValueError(f"Riferimento cella non valido: {ref}")

    # Colonna: A=0, B=1, ..., Z=25, AA=26, ...
    col = 0
    for ch in col_str:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    col -= 1  # 0-indexed

    row = int(row_str) - 1  # 0-indexed
    return row, col


def parse_range(range_str: str) -> tuple[int, int, int, int]:
    """Converte range A1:D10 -> (min_row, min_col, max_row, max_col) 0-indexed."""
    parts = range_str.split(":")
    if len(parts) != 2:
        raise ValueError(f"Range non valido: {range_str} (usa formato A1:D10)")

    r1, c1 = parse_cell_ref(parts[0])
    r2, c2 = parse_cell_ref(parts[1])
    return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)


def read_excel(file_path: str, sheet_name: str = "", cell_range: str = "",
               max_rows: int = 100, header_row: int = 1,
               no_header: bool = False) -> tuple[list[str], list[list]]:
    """
    Legge dati da un file Excel.

    Returns:
        (headers, rows) dove headers è lista di stringhe e rows è lista di liste.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERRORE: openpyxl non installato. Installa con: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(file_path, read_only=True, data_only=True)

    # Seleziona il foglio
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            print(f"ERRORE: Foglio '{sheet_name}' non trovato. Disponibili: {available}", file=sys.stderr)
            wb.close()
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Leggi tutte le righe
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([cell if cell is not None else "" for cell in row])

    wb.close()

    if not all_rows:
        return [], []

    # Applica filtro range se specificato
    if cell_range:
        min_r, min_c, max_r, max_c = parse_range(cell_range)
        filtered = []
        for r_idx, row in enumerate(all_rows):
            if min_r <= r_idx <= max_r:
                filtered_row = []
                for c_idx in range(min_c, max_c + 1):
                    if c_idx < len(row):
                        filtered_row.append(row[c_idx])
                    else:
                        filtered_row.append("")
                filtered.append(filtered_row)
        all_rows = filtered

    # Gestisci intestazione
    if no_header or not all_rows:
        # Genera intestazioni numeriche
        max_cols = max(len(r) for r in all_rows) if all_rows else 0
        headers = [f"Col_{i+1}" for i in range(max_cols)]
        data_rows = all_rows
    else:
        # Usa la riga header_row (1-indexed in input, 0-indexed internamente)
        h_idx = header_row - 1
        if cell_range:
            h_idx = 0  # se c'è un range, la prima riga del range è l'header
        if h_idx < len(all_rows):
            headers = [str(h) if h != "" else f"Col_{i+1}" for i, h in enumerate(all_rows[h_idx])]
            data_rows = all_rows[h_idx + 1:]
        else:
            max_cols = max(len(r) for r in all_rows) if all_rows else 0
            headers = [f"Col_{i+1}" for i in range(max_cols)]
            data_rows = all_rows

    # Limita righe
    truncated = False
    total_rows = len(data_rows)
    if len(data_rows) > max_rows:
        data_rows = data_rows[:max_rows]
        truncated = True

    # Normalizza larghezza (tutte le righe stessa lunghezza)
    max_cols = len(headers)
    for row in data_rows:
        while len(row) < max_cols:
            row.append("")

    if truncated:
        data_rows.append([f"... ({total_rows - max_rows} righe omesse, totale: {total_rows})"]
                         + [""] * (max_cols - 1))

    return headers, data_rows


def format_markdown(headers: list[str], rows: list[list]) -> str:
    """Formatta in tabella markdown."""
    if not headers:
        return "(Foglio vuoto)"

    # Converti tutto a stringa
    str_headers = [str(h) for h in headers]
    str_rows = [[str(cell) for cell in row] for row in rows]

    # Calcola larghezza colonne
    widths = [len(h) for h in str_headers]
    for row in str_rows:
        for i, cell in enumerate(row):
            if i < len(widths):
                widths[i] = max(widths[i], len(cell))

    # Costruisci tabella
    lines = []

    # Header
    header_line = "| " + " | ".join(h.ljust(widths[i]) for i, h in enumerate(str_headers)) + " |"
    sep_line = "| " + " | ".join("-" * widths[i] for i in range(len(str_headers))) + " |"
    lines.append(header_line)
    lines.append(sep_line)

    # Righe
    for row in str_rows:
        row_line = "| " + " | ".join(
            str(row[i]).ljust(widths[i]) if i < len(row) else " " * widths[i]
            for i in range(len(str_headers))
        ) + " |"
        lines.append(row_line)

    return "\n".join(lines)


def format_csv(headers: list[str], rows: list[list]) -> str:
    """Formatta in CSV."""
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def format_json(headers: list[str], rows: list[list]) -> str:
    """Formatta in JSON (lista di oggetti)."""
    result = []
    for row in rows:
        obj = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else ""
            # Converti tipi per JSON
            if isinstance(val, (int, float, bool)):
                obj[str(h)] = val
            else:
                obj[str(h)] = str(val) if val != "" else None
        result.append(obj)
    return json.dumps(result, ensure_ascii=False, indent=2)


def list_sheets(file_path: str) -> str:
    """Elenca i fogli di un file Excel."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "ERRORE: openpyxl non installato. Installa con: pip install openpyxl"

    wb = load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    lines = [f"Fogli nel file '{os.path.basename(file_path)}':"]
    for i, name in enumerate(sheets, 1):
        lines.append(f"  {i}. {name}")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Leggi dati da file Excel (.xlsx/.xls)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempi:
  python3 excel_reader.py dati.xlsx
  python3 excel_reader.py dati.xlsx --list-sheets
  python3 excel_reader.py dati.xlsx --sheet "Foglio2" --range A1:D20
  python3 excel_reader.py dati.xlsx --format json --max-rows 50
  python3 excel_reader.py dati.xlsx --format csv > output.csv
        """
    )
    parser.add_argument("file", help="Percorso del file Excel")
    parser.add_argument("--sheet", default="", help="Nome del foglio (default: primo)")
    parser.add_argument("--range", default="", dest="cell_range",
                        help="Range celle es. A1:D10 (default: tutto)")
    parser.add_argument("--format", default="markdown", choices=["markdown", "csv", "json"],
                        dest="fmt", help="Formato output (default: markdown)")
    parser.add_argument("--max-rows", type=int, default=100,
                        help="Max righe da mostrare (default: 100)")
    parser.add_argument("--list-sheets", action="store_true",
                        help="Elenca i fogli disponibili")
    parser.add_argument("--header", type=int, default=1,
                        help="Riga intestazione, 1-indexed (default: 1)")
    parser.add_argument("--no-header", action="store_true",
                        help="Non usare intestazione")

    args = parser.parse_args()

    # Verifica che il file esista
    if not os.path.isfile(args.file):
        print(f"ERRORE: File non trovato: {args.file}", file=sys.stderr)
        sys.exit(1)

    # Elenca fogli
    if args.list_sheets:
        print(list_sheets(args.file))
        return

    # Leggi dati
    try:
        headers, rows = read_excel(
            args.file,
            sheet_name=args.sheet,
            cell_range=args.cell_range,
            max_rows=args.max_rows,
            header_row=args.header,
            no_header=args.no_header,
        )
    except Exception as e:
        print(f"ERRORE durante la lettura: {e}", file=sys.stderr)
        sys.exit(1)

    # Formatta output
    formatters = {
        "markdown": format_markdown,
        "csv": format_csv,
        "json": format_json,
    }
    output = formatters[args.fmt](headers, rows)

    # Tronca se troppo grande
    if len(output.encode("utf-8")) > MAX_OUTPUT_BYTES:
        output = output[:MAX_OUTPUT_BYTES] + "\n\n... (output troncato, troppo grande)"

    print(output)


if __name__ == "__main__":
    main()
